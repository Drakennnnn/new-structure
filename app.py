import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import re
import base64
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import zipfile
from docxtpl import DocxTemplate
import plotly.express as px
import plotly.graph_objects as go

# Set page configuration
st.set_page_config(
    page_title="Real Estate Cost Sheet Generator",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Define CSS styles
st.markdown("""
<style>
    .main-header {
        font-size: 32px;
        font-weight: bold;
        color: #1E40AF;
        margin-bottom: 20px;
        text-align: center;
        padding: 15px;
        border-radius: 8px;
        background-color: #EFF6FF;
        border-left: 5px solid #3B82F6;
    }
    .section-header {
        font-size: 24px;
        font-weight: bold;
        color: #2563EB;
        margin-top: 30px;
        margin-bottom: 15px;
        padding-bottom: 5px;
        border-bottom: 2px solid #BFDBFE;
    }
    .subsection-header {
        font-size: 20px;
        font-weight: bold;
        color: #3B82F6;
        margin-top: 20px;
        margin-bottom: 10px;
    }
    .info-text {
        font-size: 16px;
        color: #1F2937;
        margin-bottom: 20px;
    }
    .success-box {
        padding: 15px;
        border-radius: 5px;
        background-color: #ECFDF5;
        border-left: 5px solid #10B981;
        margin-bottom: 20px;
    }
    .warning-box {
        padding: 15px;
        border-radius: 5px;
        background-color: #FFFBEB;
        border-left: 5px solid #F59E0B;
        margin-bottom: 20px;
    }
    .error-box {
        padding: 15px;
        border-radius: 5px;
        background-color: #FEF2F2;
        border-left: 5px solid #EF4444;
        margin-bottom: 20px;
    }
    .info-box {
        padding: 15px;
        border-radius: 5px;
        background-color: #EFF6FF;
        border-left: 5px solid #3B82F6;
        margin-bottom: 20px;
    }
    .status-badge {
        display: inline-block;
        padding: 4px 10px;
        border-radius: 12px;
        font-size: 14px;
        font-weight: bold;
        margin-right: 5px;
    }
    .status-verified {
        background-color: #D1FAE5;
        color: #047857;
    }
    .status-warning {
        background-color: #FEF3C7;
        color: #B45309;
    }
    .status-error {
        background-color: #FEE2E2;
        color: #B91C1C;
    }
    .metric-card {
        padding: 15px;
        border-radius: 8px;
        background-color: #F9FAFB;
        border: 1px solid #E5E7EB;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .metric-value {
        font-size: 28px;
        font-weight: bold;
        color: #1E40AF;
    }
    .metric-label {
        font-size: 14px;
        color: #4B5563;
    }
    .metric-trend {
        font-size: 14px;
        padding-top: 5px;
    }
    .metric-trend-up {
        color: #047857;
    }
    .metric-trend-down {
        color: #B91C1C;
    }
    .footer {
        margin-top: 50px;
        text-align: center;
        color: #6B7280;
        font-size: 14px;
        padding: 20px;
        border-top: 1px solid #E5E7EB;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #F3F4F6;
        border-radius: 4px 4px 0px 0px;
        gap: 1px;
        padding-top: 10px;
        padding-bottom: 10px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #DBEAFE;
        border-radius: 4px 4px 0px 0px;
    }
    .dashboard-card {
        padding: 20px;
        border-radius: 8px;
        background-color: white;
        border: 1px solid #E5E7EB;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    .stDataFrame {
        border-radius: 8px;
        border: 1px solid #E5E7EB;
    }
    .stProgress > div > div {
        background-color: #DBEAFE;
    }
    div[data-testid="stExpander"] {
        border-radius: 8px;
        border: 1px solid #E5E7EB;
    }
    div[data-testid="stExpander"] > div[role="button"] {
        background-color: #F3F4F6;
    }
    div[data-testid="stExpander"] > div[role="button"]:hover {
        background-color: #DBEAFE;
    }
    .download-btn {
        display: inline-block;
        background-color: #3B82F6;
        color: white;
        padding: 8px 16px;
        text-decoration: none;
        border-radius: 4px;
        font-weight: bold;
        margin: 10px 0;
    }
    .download-btn:hover {
        background-color: #2563EB;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown('<div class="main-header">Real Estate Cost Sheet Generator</div>', unsafe_allow_html=True)

# Initialize session state for storing data
if 'sales_master_df' not in st.session_state:
    st.session_state.sales_master_df = None
if 'collection_df' not in st.session_state:
    st.session_state.collection_df = None
if 'accounts_info' not in st.session_state:
    st.session_state.accounts_info = None
if 'verification_results' not in st.session_state:
    st.session_state.verification_results = {}
if 'selected_customers' not in st.session_state:
    st.session_state.selected_customers = []
if 'noc_template' not in st.session_state:
    st.session_state.noc_template = None
if 'preview_data' not in st.session_state:
    st.session_state.preview_data = {}
if 'processing_log' not in st.session_state:
    st.session_state.processing_log = []
if 'dashboard_data' not in st.session_state:
    st.session_state.dashboard_data = {}
if 'active_tab' not in st.session_state:
    st.session_state.active_tab = "Upload"

# Helper functions
def log_process(message, level="info"):
    """Add a log message to the processing log"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    st.session_state.processing_log.append({
        "timestamp": timestamp,
        "message": message,
        "level": level
    })

def extract_excel_date(excel_date):
    """Convert Excel date number to Python datetime"""
    try:
        # Excel date starts from 1900-01-01
        if isinstance(excel_date, (int, float)) and excel_date > 0:
            # Handle both date formats
            try:
                return pd.to_datetime('1899-12-30') + pd.Timedelta(days=int(excel_date))
            except:
                return pd.to_datetime('1899-12-30') + pd.Timedelta(days=excel_date)
        elif isinstance(excel_date, str):
            # Try multiple date format parsing approaches
            try:
                # First try to parse directly with pandas
                parsed_date = pd.to_datetime(excel_date, errors='coerce')
                if pd.notna(parsed_date):
                    return parsed_date
                
                # Try common date formats
                for fmt in ["%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%m/%d/%Y"]:
                    try:
                        return datetime.strptime(excel_date, fmt)
                    except:
                        pass
                        
                # If all else fails, try to extract date parts
                date_pattern = r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})'
                match = re.search(date_pattern, excel_date)
                if match:
                    day, month, year = map(int, match.groups())
                    if year < 100:
                        year += 2000  # Assume 20xx for two-digit years
                    return datetime(year, month, day)
                
                # Fall back to extracting numbers
                numbers = re.findall(r'\d+', excel_date)
                if len(numbers) >= 3:
                    day = int(numbers[0])
                    month = int(numbers[1])
                    year = int(numbers[2])
                    if year < 100:
                        year += 2000
                    return datetime(year, month, day)
                    
                return None
            except:
                return None
        else:
            return excel_date
    except:
        return None

def normalize_unit_number(unit_number):
    """Normalize unit number for consistent comparison"""
    if not unit_number:
        return ""
        
    # Convert to string and clean up
    unit_str = str(unit_number).strip().upper()
    
    # Remove all spaces and ensure consistent CA prefix
    unit_str = unit_str.replace(" ", "")
    
    # Ensure consistent hyphen usage
    if "-" not in unit_str and len(unit_str) >= 5:
        # Try to add hyphen in the right place (e.g., CA071208 -> CA07-1208)
        if unit_str.startswith("CA") and unit_str[2:].isdigit():
            if len(unit_str) >= 7:  # For CA + 2-digit tower + 4-digit unit
                unit_str = unit_str[:4] + "-" + unit_str[4:]
            else:  # For CA + 1-digit tower + 3-digit unit
                unit_str = unit_str[:3] + "-" + unit_str[3:]
    
    return unit_str

def identify_sales_master_sheet(workbook):
    """Return the Annex - Sales Master sheet"""
    # Check if the exact sheet name exists
    if "Annex - Sales Master" in workbook.sheetnames:
        return "Annex - Sales Master"
    
    # Fallback: try to find a sheet with a similar name
    for sheet_name in workbook.sheetnames:
        if "Annex" in sheet_name and "Sales" in sheet_name:
            return sheet_name
            
    # Look for sheets with similar columns if name doesn't match
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=3):
            row_text = " ".join([str(cell.value) for cell in row if cell.value])
            if "Unit Number" in row_text and "Name of Customer" in row_text:
                return sheet_name
    
    # If still not found, return first sheet as last resort
    return workbook.sheetnames[0] if workbook.sheetnames else None

def identify_collection_sheet(workbook):
    """Return the Main Collection AC P1_P2_P3 sheet"""
    # Check if the exact sheet name exists
    if "Main Collection AC P1_P2_P3" in workbook.sheetnames:
        return "Main Collection AC P1_P2_P3"
    
    # Fallback: try to find a sheet with a similar name
    for sheet_name in workbook.sheetnames:
        if "Main Collection" in sheet_name:
            return sheet_name
            
    # Look for sheets with phase headers
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=10):
            row_text = " ".join([str(cell.value) for cell in row if cell.value])
            if "Main Collection Escrow A/c Phase" in row_text:
                return sheet_name
    
    # If still not found, return None
    return None

def parse_sales_master(sheet):
    """Parse the Annex - Sales Master sheet and return a DataFrame"""
    # Find header row - look for key column names
    header_row = None
    header_cols = {}
    
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        col_matches = 0
        for c_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                cell_lower = cell.lower()
                if any(key in cell_lower for key in ['sr no', 'customer', 'unit', 'tower', 'booking']):
                    col_matches += 1
        
        if col_matches >= 3:  # Found at least 3 key columns
            header_row = r_idx + 1
            break
    
    if header_row is None:
        log_process("Could not find header row in Sales Master sheet", "error")
        return None
    
    # Map column indices to expected column names
    header_values = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    
    # Map common column variations to standardized names
    column_mapping = {}
    for idx, header in enumerate(header_values):
        if header:
            header_str = str(header).lower()
            
            # Map variations to standard names
            if any(term in header_str for term in ['sr', 'serial']):
                column_mapping[idx] = 'Sr No'
            elif 'customer' in header_str or 'name' in header_str:
                column_mapping[idx] = 'Name of Customer'
            elif 'unit' in header_str and 'number' in header_str:
                column_mapping[idx] = 'Unit Number'
            elif 'tower' in header_str:
                column_mapping[idx] = 'Tower No'
            elif 'booking' in header_str and 'date' in header_str:
                column_mapping[idx] = 'Booking date'
            elif 'status' in header_str and ('booking' in header_str or 'active' in header_str or 'cancelled' in header_str):
                column_mapping[idx] = 'Booking Status'
            elif ('self' in header_str and 'fund' in header_str) or ('loan' in header_str and 'avail' in header_str):
                column_mapping[idx] = 'Self-funded or loan availed'
            elif 'payment' in header_str and 'plan' in header_str:
                column_mapping[idx] = 'Payment Plan'
            elif 'builder' in header_str and 'buyer' in header_str and 'agreement' in header_str:
                column_mapping[idx] = 'Builder Buyer Agreement'
            elif 'type' in header_str and 'unit' in header_str:
                column_mapping[idx] = 'Type of Unit'
            elif 'area' in header_str and not 'carpet' in header_str:
                column_mapping[idx] = 'Area(sqft)'
            elif 'carpet' in header_str and 'area' in header_str:
                column_mapping[idx] = 'Carpet Area(sqft)'
            elif 'bsp' in header_str and 'sqft' in header_str:
                column_mapping[idx] = 'BSP/SqFt'
            elif ('basic' in header_str and 'price' in header_str) or ('basic' in header_str and 'exl' in header_str):
                column_mapping[idx] = 'Basic Price ( Exl Taxes)'
            elif 'received' in header_str and ('amount' in header_str or 'amt' in header_str) and not 'tax' in header_str and not ('inc' in header_str):
                column_mapping[idx] = 'Amount received ( Exl Taxes)'
            elif 'tax' in header_str and 'received' in header_str:
                column_mapping[idx] = 'Taxes Received'
            elif 'received' in header_str and ('inc' in header_str or 'with' in header_str) and 'tax' in header_str:
                column_mapping[idx] = 'Amount received (Inc Taxes)'
            elif ('balance' in header_str and 'receivable' in header_str) or ('total' in header_str and 'consideration' in header_str and 'balance' in header_str):
                column_mapping[idx] = 'Balance receivables (Total Sale Consideration )'
            elif 'broker' in header_str and 'name' in header_str:
                column_mapping[idx] = 'Broker Name'
            elif 'co' in header_str and 'applicant' in header_str:
                column_mapping[idx] = 'CO-APPLICANT NAME'
            else:
                # Keep original header for other columns
                column_mapping[idx] = header
    
    # Log the column mapping for debugging
    log_process(f"Found {len(column_mapping)} columns in Sales Master sheet")
    
    # Read data into a DataFrame
    data = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip completely empty rows
        if all(cell is None or cell == '' for cell in row):
            continue
        
        row_data = {}
        for idx, cell in enumerate(row):
            if idx in column_mapping:
                row_data[column_mapping[idx]] = cell
        
        # Only add rows with unit number or customer name
        if ('Unit Number' in row_data and row_data['Unit Number']) or \
           ('Name of Customer' in row_data and row_data['Name of Customer']):
            data.append(row_data)
    
    df = pd.DataFrame(data)
    
    # Process date columns
    date_columns = [col for col in df.columns if 'date' in col.lower()]
    for col in date_columns:
        df[col] = df[col].apply(extract_excel_date)
    
    # Add normalized unit number column for matching
    if 'Unit Number' in df.columns:
        df['Normalized Unit Number'] = df['Unit Number'].apply(normalize_unit_number)
    
    log_process(f"Processed {len(df)} rows from Sales Master sheet")
    
    return df

def find_phase_transitions(sheet):
    """Find phase transitions in the collection sheet"""
    phase_transitions = []
    
    # Get sheet dimensions
    if not sheet or '!ref' not in sheet:
        log_process("Invalid sheet or missing reference range", "error")
        return phase_transitions
        
    range_ref = sheet['!ref']
    range_data = openpyxl.utils.range.range_boundaries(range_ref)
    max_row = range_data[3]  # End row
    
    # Scan for phase headers
    for r in range(min(max_row, 5000)):  # Limit to first 5000 rows for performance
        # Check first few columns for phase headers
        for c in range(3):  # Check columns A, B, C
            cell_ref = f"{openpyxl.utils.get_column_letter(c+1)}{r+1}"
            cell = sheet[cell_ref] if cell_ref in sheet else None
            
            if cell and cell.value:
                cell_value = str(cell.value)
                if "Main Collection Escrow A/c Phase" in cell_value:
                    # Look for account number in nearby cells
                    account_number = None
                    
                    # Check same row, next few columns
                    for ac in range(c, c+5):
                        if ac < 26:  # Limit to column Z
                            acct_cell_ref = f"{openpyxl.utils.get_column_letter(ac+1)}{r+1}"
                            acct_cell = sheet[acct_cell_ref] if acct_cell_ref in sheet else None
                            
                            if acct_cell and acct_cell.value and isinstance(acct_cell.value, (int, str)):
                                acct_value = str(acct_cell.value)
                                if acct_value.isdigit() and len(acct_value) >= 10:
                                    account_number = acct_value
                                    break
                    
                    # If not found, check next rows
                    if not account_number:
                        for nr in range(r+1, r+5):
                            if nr < max_row:
                                for ac in range(5):
                                    acct_cell_ref = f"{openpyxl.utils.get_column_letter(ac+1)}{nr+1}"
                                    acct_cell = sheet[acct_cell_ref] if acct_cell_ref in sheet else None
                                    
                                    if acct_cell and acct_cell.value and isinstance(acct_cell.value, (int, str)):
                                        acct_value = str(acct_cell.value)
                                        if acct_value.isdigit() and len(acct_value) >= 10:
                                            account_number = acct_value
                                            break
                            
                            if account_number:
                                break
                    
                    # Find the header row (with column names)
                    header_row = None
                    for hr in range(r, r+5):
                        if hr < max_row:
                            for hc in range(5):
                                header_cell_ref = f"{openpyxl.utils.get_column_letter(hc+1)}{hr+1}"
                                header_cell = sheet[header_cell_ref] if header_cell_ref in sheet else None
                                
                                if header_cell and header_cell.value and isinstance(header_cell.value, str):
                                    header_value = str(header_cell.value)
                                    if "Txn Date" in header_value:
                                        header_row = hr
                                        break
                            
                            if header_row is not None:
                                break
                    
                    # Add the phase transition
                    phase_transitions.append({
                        "row": r + 1,
                        "name": cell_value,
                        "account_number": account_number,
                        "header_row": header_row + 1 if header_row is not None else None
                    })
    
    # For each phase, determine the end row (start of next phase - 1)
    for i in range(len(phase_transitions)):
        if i < len(phase_transitions) - 1:
            phase_transitions[i]["end_row"] = phase_transitions[i+1]["row"] - 1
        else:
            phase_transitions[i]["end_row"] = max_row
    
    log_process(f"Found {len(phase_transitions)} phase transitions in collection sheet")
    for phase in phase_transitions:
        log_process(f"Phase: {phase['name']}, Rows: {phase['row']}-{phase['end_row']}, Account: {phase['account_number']}")
    
    return phase_transitions

def identify_account_sections(sheet):
    """Identify different account sections in the collection sheet"""
    # Find phase transitions
    phase_transitions = find_phase_transitions(sheet)
    
    # Convert to account info format
    accounts = []
    
    for phase in phase_transitions:
        if phase["header_row"]:
            # Get header indices for this phase
            header_row = phase["header_row"] - 1  # Convert to 0-based
            header_indices = {}
            
            # Map headers to column indices
            for c in range(min(25, sheet.max_column)):  # Check up to 25 columns
                cell_ref = f"{openpyxl.utils.get_column_letter(c+1)}{header_row+1}"
                cell = sheet[cell_ref] if cell_ref in sheet else None
                
                if cell and cell.value:
                    header_value = str(cell.value).lower()
                    
                    if 'txn date' in header_value:
                        header_indices['date'] = c
                    elif 'description' in header_value:
                        header_indices['description'] = c
                    elif 'amount' in header_value and 'running' not in header_value:
                        header_indices['amount'] = c
                    elif ('dr' in header_value and 'cr' in header_value) or header_value == 'dr/cr':
                        header_indices['type'] = c
                    elif 'sales' in header_value and 'tag' in header_value:
                        header_indices['sales_tag'] = c
                    elif 'running' in header_value and 'total' in header_value:
                        header_indices['running_total'] = c
            
            account = {
                'name': phase["name"],
                'number': phase["account_number"],
                'start_row': phase["row"],
                'end_row': phase["end_row"],
                'header_row': header_row,
                'header_indices': header_indices
            }
            
            accounts.append(account)
    
    return accounts

def parse_collection_transactions(sheet, accounts_info):
    """Parse transactions from collection sheet based on identified account sections"""
    all_transactions = []
    
    for account in accounts_info:
        # Get header indices for this account
        header_indices = account['header_indices']
        
        # Skip if we don't have necessary columns
        if 'date' not in header_indices or 'amount' not in header_indices:
            log_process(f"Missing required columns in account {account['name']}", "warning")
            continue
        
        # Read transactions for this account
        for r_idx in range(account['header_row'] + 1, account.get('end_row', sheet.max_row) + 1):
            # Skip rows with no data
            cell_ref = f"{openpyxl.utils.get_column_letter(header_indices['date']+1)}{r_idx+1}"
            date_cell = sheet[cell_ref] if cell_ref in sheet else None
            
            cell_ref = f"{openpyxl.utils.get_column_letter(header_indices['amount']+1)}{r_idx+1}"
            amount_cell = sheet[cell_ref] if cell_ref in sheet else None
            
            # Skip if date or amount is missing
            if not date_cell or not date_cell.value or not amount_cell or amount_cell.value is None:
                continue
                
            # Only include rows with amount (valid transactions)
            if isinstance(amount_cell.value, (int, float)) and amount_cell.value != 0:
                transaction = {
                    'account_name': account['name'],
                    'account_number': account['number'],
                    'row': r_idx + 1
                }
                
                # Extract data based on header indices
                for field, idx in header_indices.items():
                    cell_ref = f"{openpyxl.utils.get_column_letter(idx+1)}{r_idx+1}"
                    cell = sheet[cell_ref] if cell_ref in sheet else None
                    
                    if cell and cell.value is not None:
                        # Special handling for sales_tag and type fields
                        if field == 'sales_tag':
                            transaction[field] = str(cell.value) if cell.value is not None else None
                        elif field == 'type':
                            transaction[field] = str(cell.value) if cell.value is not None else None
                        else:
                            transaction[field] = cell.value
                
                # Convert Excel date to Python datetime
                if 'date' in transaction:
                    transaction['date'] = extract_excel_date(transaction['date'])
                
                # Add normalized sales tag if present
                if 'sales_tag' in transaction and transaction['sales_tag']:
                    transaction['normalized_sales_tag'] = normalize_unit_number(transaction['sales_tag'])
                
                all_transactions.append(transaction)
    
    # Convert to DataFrame
    if all_transactions:
        df = pd.DataFrame(all_transactions)
        log_process(f"Extracted {len(df)} transactions from collection sheet")
        return df
    else:
        log_process("No transactions found in collection sheet", "warning")
        return pd.DataFrame()

def match_transactions_to_units(sales_master_df, collection_df):
    """Match transactions to units using robust matching logic"""
    # Check if we have the necessary data
    if sales_master_df is None or collection_df is None:
        log_process("Missing data for transaction matching", "error")
        return {}
    
    # Check if we have the normalized unit numbers and sales tags
    if 'Normalized Unit Number' not in sales_master_df.columns:
        sales_master_df['Normalized Unit Number'] = sales_master_df['Unit Number'].apply(normalize_unit_number)
    
    if 'sales_tag' not in collection_df.columns:
        log_process("No sales_tag column in collection data", "warning")
        return {}
    
    if 'normalized_sales_tag' not in collection_df.columns:
        collection_df['normalized_sales_tag'] = collection_df['sales_tag'].apply(normalize_unit_number)
    
    # Create mapping dictionary
    unit_to_transactions = {}
    
    # For each unit in sales master
    for _, unit in sales_master_df.iterrows():
        unit_number = unit['Unit Number']
        
        if not unit_number:
            continue
            
        normalized_unit = unit['Normalized Unit Number']
        
        # Try different matching patterns
        matches = []
        
        # 1. Direct match with normalized unit/tag
        direct_matches = collection_df[collection_df['normalized_sales_tag'] == normalized_unit]
        if not direct_matches.empty:
            matches.extend(direct_matches.to_dict('records'))
        
        # 2. Try matching without CA prefix
        if normalized_unit.startswith('CA'):
            # Remove CA prefix
            no_prefix = normalized_unit[2:]
            prefix_matches = collection_df[collection_df['normalized_sales_tag'].str.contains(no_prefix, regex=False, na=False)]
            if not prefix_matches.empty:
                # Filter out false matches (where the match is a substring of a larger number)
                for _, match in prefix_matches.iterrows():
                    if match['normalized_sales_tag'] not in [m['normalized_sales_tag'] for m in matches]:
                        matches.append(match.to_dict())
        
        # 3. Try matching numeric part only
        if '-' in normalized_unit:
            # Get the numeric part after the hyphen
            numeric_part = normalized_unit.split('-')[-1]
            if numeric_part.isdigit():
                numeric_matches = collection_df[collection_df['normalized_sales_tag'].str.contains(numeric_part, regex=False, na=False)]
                if not numeric_matches.empty:
                    # Filter out false matches
                    for _, match in numeric_matches.iterrows():
                        if match['normalized_sales_tag'] not in [m['normalized_sales_tag'] for m in matches]:
                            matches.append(match.to_dict())
        
        # Store all matches for this unit
        if matches:
            unit_to_transactions[unit_number] = matches
    
    log_process(f"Matched {len(unit_to_transactions)} units with transactions")
    return unit_to_transactions

def verify_transactions(sales_master_df, collection_df):
    """Verify transactions against customer data"""
    verification_results = {}
    
    # Match transactions to units
    unit_transactions_map = match_transactions_to_units(sales_master_df, collection_df)
    
    # Process each customer
    for _, customer in sales_master_df.iterrows():
        unit_number = customer.get('Unit Number')
        customer_name = customer.get('Name of Customer')
        
        if not unit_number:
            continue
        
        # Find related transactions for this unit
        unit_transactions = unit_transactions_map.get(unit_number, [])
        
        # Calculate expected amounts
        expected_amount = customer.get('Amount received (Inc Taxes)', 0)
        if pd.isna(expected_amount):
            expected_amount = 0
        
        expected_base_amount = customer.get('Amount received ( Exl Taxes)', 0) 
        if pd.isna(expected_base_amount):
            expected_base_amount = 0
        
        expected_tax_amount = customer.get('Taxes Received', 0)
        if pd.isna(expected_tax_amount):
            expected_tax_amount = 0
        
        # Calculate actual received from transactions
        credit_transactions = [t for t in unit_transactions if t.get('type', '').upper() == 'C']
        debit_transactions = [t for t in unit_transactions if t.get('type', '').upper() == 'D']
        
        total_credits = sum(t.get('amount', 0) for t in credit_transactions)
        total_debits = sum(t.get('amount', 0) for t in debit_transactions)
        
        # Calculate net amount (credits - debits)
        actual_amount = total_credits - total_debits
        
        # Check for bounced transactions
        bounced_transactions = []
        for cr_txn in credit_transactions:
            if 'date' in cr_txn and 'amount' in cr_txn:
                cr_date = cr_txn['date']
                cr_amount = cr_txn['amount']
                
                # Look for debits with same amount within 7 days (potential bounce)
                potential_bounces = [
                    d for d in debit_transactions
                    if 'date' in d and d['date'] and
                    cr_date and isinstance(cr_date, (datetime, pd.Timestamp)) and
                    d['date'] and isinstance(d['date'], (datetime, pd.Timestamp)) and
                    d['date'] >= cr_date and 
                    d['date'] <= cr_date + timedelta(days=7) and
                    abs(d['amount'] - cr_amount) < 0.01  # Allow for minor rounding differences
                ]
                
                for bounce in potential_bounces:
                    bounced_transactions.append({
                        'credit_date': cr_date,
                        'credit_amount': cr_amount,
                        'debit_date': bounce.get('date'),
                        'debit_amount': bounce.get('amount'),
                        'description': bounce.get('description', '')
                    })
        
        # Determine verification status
        # Tolerance of 1 rupee for floating point discrepancies
        amount_match = abs(actual_amount - expected_amount) <= 1
        
        status = "verified" if amount_match and not bounced_transactions else "warning"
        
        if unit_transactions and not amount_match:
            status = "error"
        
        verification_results[unit_number] = {
            'unit_number': unit_number,
            'customer_name': customer_name,
            'expected_amount': expected_amount,
            'expected_base_amount': expected_base_amount,
            'expected_tax_amount': expected_tax_amount,
            'actual_amount': actual_amount,
            'amount_match': amount_match,
            'total_credits': total_credits,
            'total_debits': total_debits,
            'transaction_count': len(unit_transactions),
            'bounced_transactions': bounced_transactions,
            'has_bounced': len(bounced_transactions) > 0,
            'status': status,
            'transactions': unit_transactions
        }
    
    return verification_results

def generate_cost_sheet_data(customer_info, verification_info):
    """Generate data for cost sheet based on customer info and verification results"""
    unit_number = customer_info.get('Unit Number')
    if not unit_number:
        return None
    
    # Format unit number consistently
    if isinstance(unit_number, str):
        unit_number = unit_number.strip().upper()
    else:
        unit_number = str(unit_number).strip().upper()
    
    tower_no = customer_info.get('Tower No', '')
    if tower_no and not isinstance(tower_no, str):
        tower_no = str(tower_no)
    
    # Format as "CA 04-402" format
    formatted_unit = unit_number
    if '-' not in unit_number and tower_no:
        # Try to form a properly formatted unit number
        unit_part = unit_number.replace('CA', '').replace(' ', '')
        formatted_unit = f"{tower_no}-{unit_part}"
    
    formatted_unit = formatted_unit.replace(" ", "-").upper()
    
    # Get values with proper null/None checking
    super_area = customer_info.get('Area(sqft)', 0)
    if super_area is None or pd.isna(super_area):
        super_area = 0
        
    carpet_area = customer_info.get('Carpet Area(sqft)', 0)
    if carpet_area is None or pd.isna(carpet_area):
        carpet_area = 0
        
    bsp_rate = customer_info.get('BSP/SqFt', 0)
    if bsp_rate is None or pd.isna(bsp_rate):
        bsp_rate = 0
        
    bsp_amount = customer_info.get('Basic Price ( Exl Taxes)', 0)
    if bsp_amount is None or pd.isna(bsp_amount):
        bsp_amount = 0
    
    # Basic customer and unit info
    cost_sheet_data = {
        'unit_number': unit_number,
        'formatted_unit': formatted_unit,
        'tower': tower_no,
        'customer_name': customer_info.get('Name of Customer', ''),
        'booking_date': customer_info.get('Booking date'),
        'payment_plan': customer_info.get('Payment Plan', ''),
        'super_area': super_area,
        'carpet_area': carpet_area,
    }
    
    # Cost calculations
    ifms_rate = 25  # Standard rates based on your examples
    ifms_amount = ifms_rate * super_area
    amc_rate = 3
    amc_amount = amc_rate * super_area * 12  # Annual (12 months)
    gst_rate = 5  # 5% on BSP
    gst_amount = 0.05 * bsp_amount
    amc_gst_rate = 18  # 18% on AMC
    amc_gst_amount = 0.18 * amc_amount
    
    cost_sheet_data.update({
        'bsp_rate': bsp_rate,
        'bsp_amount': bsp_amount,
        'ifms_rate': ifms_rate,
        'ifms_amount': ifms_amount,
        'amc_rate': amc_rate,
        'amc_amount': amc_amount,
        'gst_rate': gst_rate,
        'gst_amount': gst_amount,
        'amc_gst_rate': amc_gst_rate,
        'amc_gst_amount': amc_gst_amount,
    })
    
    # Payment information
    verification = verification_info.get(unit_number, {})
    
    # Handle null/None values
    expected_base_amount = verification.get('expected_base_amount', 0)
    if expected_base_amount is None or pd.isna(expected_base_amount):
        expected_base_amount = 0
        
    expected_tax_amount = verification.get('expected_tax_amount', 0) 
    if expected_tax_amount is None or pd.isna(expected_tax_amount):
        expected_tax_amount = 0
        
    expected_amount = verification.get('expected_amount', 0)
    if expected_amount is None or pd.isna(expected_amount):
        expected_amount = 0
        
    balance_receivable = customer_info.get('Balance receivables (Total Sale Consideration )', 0)
    if balance_receivable is None or pd.isna(balance_receivable):
        balance_receivable = 0
    
    cost_sheet_data.update({
        'amount_received': expected_base_amount,
        'gst_received': expected_tax_amount,
        'total_received': expected_amount,
        'balance_receivable': balance_receivable,
    })
    
    # Bank transaction details
    transactions = verification.get('transactions', [])
    cost_sheet_data['transactions'] = sorted(transactions, key=lambda x: x.get('date', pd.NaT) or pd.NaT)
    
    # Calculate total consideration and balance receivable
    total_consideration = bsp_amount + ifms_amount + amc_amount
    cost_sheet_data['total_consideration'] = total_consideration
    cost_sheet_data['balance_receivable'] = total_consideration - expected_base_amount
    
    # Broker information
    broker_name = customer_info.get('Broker Name', '')
    if not broker_name or pd.isna(broker_name):
        broker_name = 'DIRECT'
        
    broker_rate = 100  # Standard rate from examples
    broker_amount = broker_rate * super_area
    
    cost_sheet_data.update({
        'broker_name': broker_name,
        'broker_rate': broker_rate,
        'broker_amount': broker_amount,
    })
    
    # Calculate floor number from unit number
    floor_number = "Ground"
    try:
        unit_parts = formatted_unit.split('-')
        if len(unit_parts) > 1:
            unit_num_part = unit_parts[-1]
            if unit_num_part.isdigit():
                floor = int(unit_num_part) // 100
                floor_number = f"{floor}th" if floor != 0 else "Ground"
    except:
        pass
    
    # Additional info
    home_loan = customer_info.get('Self-funded or loan availed', '')
    if not home_loan or pd.isna(home_loan):
        home_loan = 'N/A'
        
    co_applicant = customer_info.get('CO-APPLICANT NAME', '')
    if not co_applicant or pd.isna(co_applicant):
        co_applicant = 'N/A'
    
    cost_sheet_data.update({
        'floor_number': floor_number,
        'home_loan': home_loan,
        'co_applicant': co_applicant
    })
    
    return cost_sheet_data

def generate_cost_sheet_excel(cost_sheet_data):
    """Generate Excel file for cost sheet"""
    if not cost_sheet_data:
        return None
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create our three sheets
    wb.remove(wb.active)
    data_entry_sheet = wb.create_sheet(f"{cost_sheet_data['formatted_unit']} - Data Entry")
    bank_credit_sheet = wb.create_sheet("Bank Credit Details")
    noc_sheet = wb.create_sheet("Sales NOC SWAMIH")
    
    # ----- Data Entry Sheet -----
    
    # Title row
    data_entry_sheet['A1'] = "COST SHEET"
    data_entry_sheet['A1'].font = Font(bold=True, size=14)
    
    # Header row
    data_entry_sheet['A2'] = "Particulars"
    data_entry_sheet['B2'] = "Details"
    data_entry_sheet['C2'] = "Remarks"
    
    # Apply styling to header
    for col in ['A2', 'B2', 'C2']:
        data_entry_sheet[col].font = Font(bold=True)
        data_entry_sheet[col].alignment = Alignment(horizontal='center')
    
    # Customer info
    data_entry_sheet['A3'] = "DATE OF BOOKING             "
    data_entry_sheet['B3'] = cost_sheet_data['booking_date']
    
    data_entry_sheet['A4'] = "APPLICANT NAME                 "
    data_entry_sheet['B4'] = cost_sheet_data['customer_name']
    
    data_entry_sheet['A5'] = "CO-APPLICANT NAME          "
    data_entry_sheet['B5'] = cost_sheet_data.get('co_applicant', 'N/A')
    
    data_entry_sheet['A6'] = "PAYMENT PLAN"
    data_entry_sheet['B6'] = cost_sheet_data['payment_plan']
    
    data_entry_sheet['A7'] = "TOWER:"
    data_entry_sheet['B7'] = cost_sheet_data['tower']
    
    data_entry_sheet['A8'] = "UNIT NO:"
    data_entry_sheet['B8'] = cost_sheet_data['unit_number'].split('-')[-1] if '-' in cost_sheet_data['unit_number'] else cost_sheet_data['unit_number']
    
    data_entry_sheet['A9'] = "FLOOR NUMBER"
    data_entry_sheet['B9'] = cost_sheet_data['floor_number']
    
    data_entry_sheet['A10'] = "SUPER AREA(SQ. FT.)"
    data_entry_sheet['B10'] = cost_sheet_data['super_area']
    
    data_entry_sheet['A11'] = "CARPET AREA(SQ. FT.)"
    data_entry_sheet['B11'] = cost_sheet_data['carpet_area']
    
    # Cost breakdown header
    data_entry_sheet['A13'] = "Particulars"
    data_entry_sheet['B13'] = "Rate"
    data_entry_sheet['C13'] = "Amount"
    data_entry_sheet['D13'] = "Psft"
    
    for col in ['A13', 'B13', 'C13', 'D13']:
        data_entry_sheet[col].font = Font(bold=True)
        data_entry_sheet[col].alignment = Alignment(horizontal='center')
    
    # Cost breakdown data
    data_entry_sheet['A15'] = "BASIC SALE PRICE            "
    data_entry_sheet['B15'] = cost_sheet_data['bsp_rate']
    data_entry_sheet['C15'] = cost_sheet_data['bsp_amount']
    data_entry_sheet['D15'] = data_entry_sheet['C15'].value / data_entry_sheet['B10'].value if data_entry_sheet['B10'].value else 0
    
    data_entry_sheet['A16'] = "Less: Discount (if any)"
    data_entry_sheet['B16'] = "-"
    data_entry_sheet['C16'] = 0
    
    data_entry_sheet['A17'] = "NET Price"
    data_entry_sheet['C17'] = "=C15-C16"
    
    data_entry_sheet['A18'] = "ADD:-"
    
    data_entry_sheet['A19'] = "IFMS"
    data_entry_sheet['B19'] = cost_sheet_data['ifms_rate']
    data_entry_sheet['C19'] = "=B19*B10"
    data_entry_sheet['D19'] = "=C19/B10"
    
    data_entry_sheet['A20'] = "1 Year Annual Maintenance Charge"
    data_entry_sheet['B20'] = cost_sheet_data['amc_rate']
    data_entry_sheet['C20'] = "=(B20*B10)*12"
    data_entry_sheet['D20'] = "=C20/B10"
    
    data_entry_sheet['A21'] = "Lease Rent"
    data_entry_sheet['B21'] = 0
    data_entry_sheet['C21'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A22'] = "EEC"
    data_entry_sheet['B22'] = 0
    data_entry_sheet['C22'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A23'] = "FFC"
    data_entry_sheet['B23'] = 0
    data_entry_sheet['C23'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A24'] = "IDC"
    data_entry_sheet['B24'] = 0
    data_entry_sheet['C24'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A26'] = "Covered Car Parking       "
    data_entry_sheet['B26'] = 475000
    data_entry_sheet['C26'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A27'] = "Additional Car Parking"
    data_entry_sheet['B27'] = 475000
    data_entry_sheet['C27'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A28'] = "Club Membership           "
    data_entry_sheet['B28'] = 100000
    data_entry_sheet['C28'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A29'] = "Power Backup( 1.K.V.A)                  "
    data_entry_sheet['B29'] = 20000
    data_entry_sheet['C29'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A30'] = "Add.Power Backup( 3.K.V.A)                  "
    data_entry_sheet['B30'] = 60000
    data_entry_sheet['C30'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A31'] = "ADD:               "
    
    data_entry_sheet['A32'] = "PLC'S:-"
    
    data_entry_sheet['A33'] = "PARK"
    data_entry_sheet['B33'] = 150
    data_entry_sheet['C33'] = "INCLUSIVE IN BSP"
    
    data_entry_sheet['A34'] = "CLUB"
    data_entry_sheet['B34'] = 100
    data_entry_sheet['C34'] = "NA"
    
    data_entry_sheet['A35'] = "CORNER"
    data_entry_sheet['B35'] = 100
    data_entry_sheet['C35'] = "NA"
    
    data_entry_sheet['A36'] = "ROAD"
    data_entry_sheet['B36'] = 100
    data_entry_sheet['C36'] = "NA"
    
    # Total considerations
    data_entry_sheet['A38'] = "Total Sale Consideration "
    data_entry_sheet['C38'] = "=SUM(C17:C37)"
    
    data_entry_sheet['A39'] = "GST AMOUNT "
    data_entry_sheet['C39'] = "=(C15)*5%"
    data_entry_sheet['D39'] = "5% GST"
    
    data_entry_sheet['A40'] = "AMC GST"
    data_entry_sheet['C40'] = "=C20*18%"
    data_entry_sheet['D40'] = "18% GST"
    
    data_entry_sheet['A41'] = "Total GST"
    data_entry_sheet['C41'] = "=C40+C39"
    
    data_entry_sheet['A42'] = "GRAND TOTAL"
    data_entry_sheet['C42'] = "=C41+C38"
    
    # Receipt details
    data_entry_sheet['A44'] = "Receipt details"
    
    data_entry_sheet['A45'] = "Amount Received"
    data_entry_sheet['C45'] = cost_sheet_data['amount_received']
    
    data_entry_sheet['A46'] = "BALANCE RECEIVABLE"
    data_entry_sheet['C46'] = "=C38-C45"
    
    data_entry_sheet['A47'] = "GST received"
    data_entry_sheet['C47'] = cost_sheet_data['gst_received']
    
    data_entry_sheet['A48'] = "BALANCE GST RECEIVABLE"
    data_entry_sheet['C48'] = "=C41-C47"
    
    # Broker info
    data_entry_sheet['A50'] = "Brokerage"
    data_entry_sheet['B50'] = cost_sheet_data['broker_rate']
    data_entry_sheet['C50'] = "=B50*B10"
    
    data_entry_sheet['A51'] = "Broker Name "
    data_entry_sheet['C51'] = cost_sheet_data['broker_name']
    
    data_entry_sheet['A52'] = "Amount to be collected Excluding Brokerage"
    data_entry_sheet['C52'] = "=C42-C50"
    
    # Additional information
    data_entry_sheet['A54'] = "Home Loan Taken from"
    data_entry_sheet['C54'] = cost_sheet_data.get('home_loan', 'N/A')
    
    data_entry_sheet['A55'] = "Contact Number of Purchaser"
    data_entry_sheet['C55'] = ""  # Would need to add this to the sales master parsing
    
    data_entry_sheet['A56'] = "Address of Purchaser"
    data_entry_sheet['C56'] = ""  # Would need to add this to the sales master parsing
    
    data_entry_sheet['A58'] = "Sale Status as per DTD"
    data_entry_sheet['C58'] = "Unsold"
    
    data_entry_sheet['A59'] = "Sale Status as per Actual "
    data_entry_sheet['C59'] = "Sold"

    # Adjust column widths
    for col, width in [('A', 40), ('B', 20), ('C', 25), ('D', 15)]:
        data_entry_sheet.column_dimensions[col].width = width
    
    # ----- Bank Credit Details Sheet -----
    bank_credit_sheet['A1'] = "Date"
    bank_credit_sheet['B1'] = "Amount received"
    bank_credit_sheet['C1'] = "Bank A/c Name"
    bank_credit_sheet['D1'] = "Bank A/c No"
    bank_credit_sheet['E1'] = "Bank statement verified"
    
    # Apply styling to header
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        bank_credit_sheet[col].font = Font(bold=True)
        bank_credit_sheet[col].alignment = Alignment(horizontal='center')
    
    # Add transaction data
    transactions = cost_sheet_data.get('transactions', [])
    credit_transactions = [t for t in transactions if t.get('type') == 'C']
    
    for i, txn in enumerate(credit_transactions, start=2):
        bank_credit_sheet[f'A{i}'] = txn.get('date')
        bank_credit_sheet[f'B{i}'] = txn.get('amount')
        bank_credit_sheet[f'C{i}'] = txn.get('account_name')
        bank_credit_sheet[f'D{i}'] = txn.get('account_number')
        bank_credit_sheet[f'E{i}'] = "Yes"
    
    # Add total row
    total_row = len(credit_transactions) + 2
    bank_credit_sheet[f'A{total_row}'] = "Total Sum received"
    bank_credit_sheet[f'B{total_row}'] = f"=SUM(B2:B{total_row-1})"
    
    # Adjust column widths
    for col, width in [('A', 15), ('B', 15), ('C', 35), ('D', 20), ('E', 20)]:
        bank_credit_sheet.column_dimensions[col].width = width
    
    # ----- Sales NOC SWAMIH Sheet -----
    noc_sheet['B1'] = "Particulars"
    noc_sheet['C1'] = "Details"
    
    noc_sheet['B2'] = "Flat/Unit No."
    noc_sheet['C2'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B8"
    
    noc_sheet['B3'] = "Floor No."
    noc_sheet['C3'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B9"
    
    noc_sheet['B4'] = "Building Name"
    noc_sheet['C4'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B7"
    
    noc_sheet['B5'] = "Carpet Area of the Flat / Unit (Sq.Ft.)"
    noc_sheet['C5'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B11"
    
    noc_sheet['B6'] = "Name of the Applicant (Purchaser)"
    noc_sheet['C6'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B4"
    
    noc_sheet['B7'] = "Name of the Co-Applicant (Co- Purchaser)"
    noc_sheet['C7'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B5"
    
    noc_sheet['B8'] = "Total Sale Consideration (including parking charges) (Excl. GST)"
    noc_sheet['C8'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C38"
    
    noc_sheet['B9'] = "Total GST"
    noc_sheet['C9'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C41"
    
    noc_sheet['B10'] = "Sale Consideration Amount Received as on date (Excl. GST)"
    noc_sheet['C10'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C45"
    
    noc_sheet['B11'] = "GST Amount received as on date"
    noc_sheet['C11'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C47"
    
    noc_sheet['B12'] = "Balance Amount yet to be received (excluding taxes)"
    noc_sheet['C12'] = "=C8-C10"
    
    noc_sheet['B13'] = "Booking Date"
    noc_sheet['C13'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!B3"
    
    noc_sheet['B14'] = "Home loan taken from"
    noc_sheet['C14'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C54"
    
    noc_sheet['B15'] = "Contact number of Purchaser"
    noc_sheet['C15'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C55"
    
    noc_sheet['B16'] = "Address of Homebuyer of Purchaser"
    noc_sheet['C16'] = f"='{cost_sheet_data['formatted_unit']} - Data Entry'!C56"
    
    # Adjust column widths
    for col, width in [('B', 60), ('C', 30)]:
        noc_sheet.column_dimensions[col].width = width
    
    # Save to BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def generate_noc_document(customer_info, template_file):
    """Generate NOC document for a customer"""
    if not template_file or not customer_info:
        return None
    
    try:
        # Load the template
        doc = DocxTemplate(template_file)
        
        # Format unit number consistently
        unit_number = customer_info.get('Unit Number')
        if isinstance(unit_number, str):
            unit_number = unit_number.strip().upper()
        else:
            unit_number = str(unit_number).strip().upper()
        
        tower_no = customer_info.get('Tower No', '')
        if tower_no and not isinstance(tower_no, str):
            tower_no = str(tower_no)
        
        # Prepare context data for template
        context = {
            'unit_no': unit_number,
            'floor_no': f"{int(unit_number.split('-')[-1]) // 100}th" if '-' in unit_number else "Ground",
            'building_name': tower_no,
            'booking_date': customer_info.get('Booking date').strftime('%d-%m-%Y') if customer_info.get('Booking date') else 'N/A',
            'saleable_area': customer_info.get('Area(sqft)'),
            'carpet_area': customer_info.get('Carpet Area(sqft)'),
            'applicant_name': customer_info.get('Name of Customer'),
            'co_applicant_name': customer_info.get('CO-APPLICANT NAME', 'N/A'),
            'today_date': datetime.now().strftime('%d-%m-%Y')
        }
        
        # Render the template with the data
        doc.render(context)
        
        # Save to a byte stream
        output = io.BytesIO()
        doc.save(output)
        output.seek(0)
        
        return output
    
    except Exception as e:
        log_process(f"Error generating NOC document: {str(e)}", "error")
        return None

def create_download_link(file_obj, file_name):
    """Create a download link for a file object"""
    b64 = base64.b64encode(file_obj.read()).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{file_name}" class="download-btn">{file_name}</a>'
    return href

def calculate_dashboard_data(sales_master_df, verification_results):
    """Calculate statistics for the dashboard"""
    dashboard_data = {}
    
    if sales_master_df is None or verification_results is None:
        return dashboard_data
    
    # Get total units and units with transactions
    total_units = len(sales_master_df)
    units_with_transactions = sum(1 for v in verification_results.values() if v['transaction_count'] > 0)
    
    # Calculate completion percentages
    unit_completion = []
    for unit, verification in verification_results.items():
        # Get total consideration and amount received
        total_consideration = 0
        amount_received = 0
        
        # Try to get from verification
        if 'expected_amount' in verification:
            amount_received = verification['expected_amount']
        
        # Try to get from sales_master_df
        unit_data = sales_master_df[sales_master_df['Unit Number'] == unit]
        if not unit_data.empty:
            # Get total consideration
            total_consideration_col = 'Total \r\nConsideration ( Exl Taxes)\r\n'
            if total_consideration_col in unit_data.columns:
                total_consideration = unit_data[total_consideration_col].iloc[0]
            elif 'Basic Price ( Exl Taxes)' in unit_data.columns:
                total_consideration = unit_data['Basic Price ( Exl Taxes)'].iloc[0]
            
            # If not in verification, get from sales_master_df
            if amount_received == 0 and 'Amount received (Inc Taxes)' in unit_data.columns:
                amount_received = unit_data['Amount received (Inc Taxes)'].iloc[0]
        
        if total_consideration and total_consideration > 0:
            completion_pct = min(100, (amount_received / total_consideration) * 100)
        else:
            completion_pct = 0
            
        customer_name = verification['customer_name']
        
        unit_completion.append({
            'unit': unit,
            'customer_name': customer_name,
            'total_consideration': total_consideration,
            'amount_received': amount_received,
            'completion_pct': completion_pct,
            'status': verification['status']
        })
    
    # Tower-wise statistics
    tower_stats = {}
    for _, row in sales_master_df.iterrows():
        tower = row.get('Tower No', 'Unknown')
        if not tower or pd.isna(tower):
            tower = 'Unknown'
            
        tower = str(tower)
        
        if tower not in tower_stats:
            tower_stats[tower] = {
                'total_units': 0,
                'active_units': 0,
                'total_consideration': 0,
                'amount_received': 0
            }
            
        tower_stats[tower]['total_units'] += 1
        
        # Check if unit is active
        booking_status = row.get('Booking Status')
        if booking_status and 'active' in str(booking_status).lower():
            tower_stats[tower]['active_units'] += 1
            
        # Add financial data
        total_consideration = row.get('Total \r\nConsideration ( Exl Taxes)\r\n', 0)
        if pd.isna(total_consideration) or total_consideration == 0:
            total_consideration = row.get('Basic Price ( Exl Taxes)', 0)
            
        amount_received = row.get('Amount received (Inc Taxes)', 0)
        
        if not pd.isna(total_consideration):
            tower_stats[tower]['total_consideration'] += total_consideration
            
        if not pd.isna(amount_received):
            tower_stats[tower]['amount_received'] += amount_received
    
    # Calculate overall statistics
    total_consideration = sum(stats['total_consideration'] for stats in tower_stats.values())
    total_received = sum(stats['amount_received'] for stats in tower_stats.values())
    overall_completion = (total_received / total_consideration * 100) if total_consideration > 0 else 0
    
    # Payment plan distribution
    payment_plan_stats = {}
    for _, row in sales_master_df.iterrows():
        payment_plan = row.get('Payment Plan', 'Unknown')
        if not payment_plan or pd.isna(payment_plan):
            payment_plan = 'Unknown'
            
        payment_plan = str(payment_plan)
        
        if payment_plan not in payment_plan_stats:
            payment_plan_stats[payment_plan] = {
                'count': 0,
                'total_consideration': 0,
                'amount_received': 0
            }
            
        payment_plan_stats[payment_plan]['count'] += 1
        
        # Add financial data
        total_consideration = row.get('Total \r\nConsideration ( Exl Taxes)\r\n', 0)
        if pd.isna(total_consideration) or total_consideration == 0:
            total_consideration = row.get('Basic Price ( Exl Taxes)', 0)
            
        amount_received = row.get('Amount received (Inc Taxes)', 0)
        
        if not pd.isna(total_consideration):
            payment_plan_stats[payment_plan]['total_consideration'] += total_consideration
            
        if not pd.isna(amount_received):
            payment_plan_stats[payment_plan]['amount_received'] += amount_received
    
    # Compile all data
    dashboard_data = {
        'total_units': total_units,
        'units_with_transactions': units_with_transactions,
        'unit_completion': unit_completion,
        'tower_stats': tower_stats,
        'payment_plan_stats': payment_plan_stats,
        'overall_completion': overall_completion,
        'total_consideration': total_consideration,
        'total_received': total_received
    }
    
    return dashboard_data

# Sidebar for uploading files
with st.sidebar:
    st.markdown('<div class="section-header">File Upload</div>', unsafe_allow_html=True)
    
    uploaded_sales_mis = st.file_uploader("Upload Sales MIS Template Excel", type=["xlsx", "xls"])
    
    st.markdown('<div class="section-header">Optional</div>', unsafe_allow_html=True)
    uploaded_noc_template = st.file_uploader("Upload NOC Document Template (Optional)", type=["docx"])
    
    if uploaded_noc_template:
        st.session_state.noc_template = uploaded_noc_template
    
    # Navigation
    st.markdown('<div class="section-header">Navigation</div>', unsafe_allow_html=True)
    
    if st.button("Upload & Process", key="nav_upload", use_container_width=True):
        st.session_state.active_tab = "Upload"
        
    if st.button("Customer Selection", key="nav_customers", use_container_width=True):
        st.session_state.active_tab = "Customers"
        
    if st.button("Dashboard", key="nav_dashboard", use_container_width=True):
        st.session_state.active_tab = "Dashboard"
        
    if st.button("Generate Cost Sheets", key="nav_generate", use_container_width=True):
        st.session_state.active_tab = "Generate"
    
    # Display processing logs
    with st.expander("Processing Logs", expanded=False):
        for log in st.session_state.processing_log:
            if log["level"] == "error":
                st.error(log["message"])
            elif log["level"] == "warning":
                st.warning(log["message"])
            else:
                st.info(log["message"])
                
# Main area based on active tab
if st.session_state.active_tab == "Upload":
    st.markdown('<div class="section-header">Data Processing & Verification</div>', unsafe_allow_html=True)
    
    if uploaded_sales_mis:
        # Process the uploaded file
        with st.spinner('Processing Sales MIS Template...'):
            try:
                # Load workbook
                workbook = openpyxl.load_workbook(uploaded_sales_mis, data_only=True)
                
                # Identify the relevant sheets
                sales_master_sheet_name = identify_sales_master_sheet(workbook)
                collection_sheet_name = identify_collection_sheet(workbook)
                
                if not sales_master_sheet_name:
                    st.error("Could not identify Annex - Sales Master sheet in the uploaded file.")
                    log_process("Could not identify Annex - Sales Master sheet", "error")
                elif not collection_sheet_name:
                    st.error("Could not identify Main Collection sheet in the uploaded file.")
                    log_process("Could not identify Main Collection sheet", "error")
                else:
                    st.success(f"Successfully identified sheets: '{sales_master_sheet_name}' and '{collection_sheet_name}'")
                    log_process(f"Identified sheets: '{sales_master_sheet_name}' and '{collection_sheet_name}'")
                    
                    # Parse Sales Master sheet
                    sales_master_df = parse_sales_master(workbook[sales_master_sheet_name])
                    st.session_state.sales_master_df = sales_master_df
                    
                    # Identify account sections in collection sheet
                    accounts_info = identify_account_sections(workbook[collection_sheet_name])
                    st.session_state.accounts_info = accounts_info
                    
                    # Parse collection transactions
                    collection_df = parse_collection_transactions(workbook[collection_sheet_name], accounts_info)
                    st.session_state.collection_df = collection_df
                    
                    # Verify transactions against customer data
                    verification_results = verify_transactions(sales_master_df, collection_df)
                    st.session_state.verification_results = verification_results
                    
                    # Calculate dashboard data
                    dashboard_data = calculate_dashboard_data(sales_master_df, verification_results)
                    st.session_state.dashboard_data = dashboard_data
                    
                    # Show summary
                    st.markdown('<div class="section-header">Verification Summary</div>', unsafe_allow_html=True)
                    
                    verified_count = sum(1 for v in verification_results.values() if v['status'] == 'verified')
                    warning_count = sum(1 for v in verification_results.values() if v['status'] == 'warning')
                    error_count = sum(1 for v in verification_results.values() if v['status'] == 'error')
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Verified Units", verified_count, f"{verified_count/len(verification_results)*100:.1f}%")
                    with col2:
                        st.metric("Units with Warnings", warning_count, f"{warning_count/len(verification_results)*100:.1f}%")
                    with col3:
                        st.metric("Units with Errors", error_count, f"{error_count/len(verification_results)*100:.1f}%")
                    
                    # Show accounts found
                    st.markdown('<div class="section-header">Bank Accounts Identified</div>', unsafe_allow_html=True)
                    
                    accounts_df = pd.DataFrame([
                        {
                            'Account Name': account['name'],
                            'Account Number': account['number'],
                            'Transaction Count': sum(1 for _, txn in collection_df.iterrows() 
                                                if txn.get('account_number') == account['number'])
                        }
                        for account in accounts_info
                    ])
                    
                    st.dataframe(accounts_df, use_container_width=True)
                    
                    # Navigate to customer selection
                    st.success("Data processed successfully! You can now proceed to Customer Selection.")
                    if st.button("Go to Customer Selection", use_container_width=True):
                        st.session_state.active_tab = "Customers"
            
            except Exception as e:
                st.error(f"Error processing file: {str(e)}")
                log_process(f"Error processing file: {str(e)}", "error")
                import traceback
                log_process(traceback.format_exc(), "error")
    else:
        st.info("Please upload the Sales MIS Template Excel file to start.")
        st.markdown("""
        <div class="info-box">
            <h3>How to Use This Application</h3>
            <p>This application helps you generate cost sheets for real estate units based on sales data and collection information.</p>
            <ol>
                <li>Upload the Sales MIS Template Excel file containing the Annex - Sales Master and Main Collection sheets.</li>
                <li>Optionally upload a NOC Document Template if you need to generate NOC documents.</li>
                <li>The application will automatically process the data and identify accounts and transactions.</li>
                <li>Select the units you want to generate cost sheets for.</li>
                <li>Generate and download the cost sheets and NOC documents.</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)

elif st.session_state.active_tab == "Customers":
    st.markdown('<div class="section-header">Customer Selection</div>', unsafe_allow_html=True)
    
    # Check if data is processed
    if st.session_state.sales_master_df is not None and st.session_state.verification_results:
        # Create a DataFrame for display
        customers_df = pd.DataFrame([
            {
                'Select': unit in st.session_state.selected_customers,
                'Unit Number': v['unit_number'],
                'Customer Name': v['customer_name'],
                'Expected Amount': v['expected_amount'],
                'Actual Amount': v['actual_amount'],
                'Difference': v['expected_amount'] - v['actual_amount'],
                'Transaction Count': v['transaction_count'],
                'Bounced Transactions': len(v['bounced_transactions']),
                'Status': v['status']
            }
            for unit, v in st.session_state.verification_results.items()
        ])
        
        # Sort by unit number
        customers_df = customers_df.sort_values('Unit Number')
        
        # Add filtering options
        st.markdown('<div class="subsection-header">Filter Customers</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        
        with col1:
            status_filter = st.selectbox(
                "Filter by Status",
                ["All", "verified", "warning", "error"]
            )
        
        with col2:
            transaction_filter = st.selectbox(
                "Filter by Transactions",
                ["All", "With Transactions", "No Transactions"]
            )
            
        with col3:
            bounced_filter = st.selectbox(
                "Filter by Bounced Transactions",
                ["All", "With Bounced", "No Bounced"]
            )
        
        # Apply filters
        filtered_df = customers_df.copy()
        
        if status_filter != "All":
            filtered_df = filtered_df[filtered_df['Status'] == status_filter]
            
        if transaction_filter != "All":
            if transaction_filter == "With Transactions":
                filtered_df = filtered_df[filtered_df['Transaction Count'] > 0]
            else:
                filtered_df = filtered_df[filtered_df['Transaction Count'] == 0]
                
        if bounced_filter != "All":
            if bounced_filter == "With Bounced":
                filtered_df = filtered_df[filtered_df['Bounced Transactions'] > 0]
            else:
                filtered_df = filtered_df[filtered_df['Bounced Transactions'] == 0]
        
        # Use st.data_editor to make it selectable
        edited_df = st.data_editor(
            filtered_df,
            column_config={
                "Select": st.column_config.CheckboxColumn(
                    "Select",
                    help="Select customer for cost sheet generation",
                    default=False,
                ),
                "Status": st.column_config.SelectboxColumn(
                    "Status",
                    help="Verification status",
                    options=["verified", "warning", "error"],
                    required=True,
                ),
                "Expected Amount": st.column_config.NumberColumn(
                    "Expected Amount",
                    format="₹ %.2f",
                ),
                "Actual Amount": st.column_config.NumberColumn(
                    "Actual Amount",
                    format="₹ %.2f",
                ),
                "Difference": st.column_config.NumberColumn(
                    "Difference",
                    format="₹ %.2f",
                )
            },
            disabled=["Unit Number", "Customer Name", "Expected Amount", "Actual Amount", 
                     "Difference", "Transaction Count", "Bounced Transactions", "Status"],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed"
        )
        
        # Store selected customers
        st.session_state.selected_customers = edited_df[edited_df['Select']]['Unit Number'].tolist()
        
        # Show selection summary
        st.markdown(f"<div class='info-box'>Selected {len(st.session_state.selected_customers)} customers for cost sheet generation</div>", unsafe_allow_html=True)
        
        # Show detailed verification for selected customers
        if st.session_state.selected_customers:
            st.markdown('<div class="section-header">Verification Details</div>', unsafe_allow_html=True)
            
            # Create tabs for each selected customer
            tabs = st.tabs([f"{unit_no}" for unit_no in st.session_state.selected_customers])
            
            for i, tab in enumerate(tabs):
                unit_no = st.session_state.selected_customers[i]
                verification = st.session_state.verification_results.get(unit_no, {})
                
                with tab:
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown('<div class="subsection-header">Customer Information</div>', unsafe_allow_html=True)
                        st.write(f"**Customer:** {verification.get('customer_name', 'N/A')}")
                        st.write(f"**Unit:** {unit_no}")
                        
                        customer_row = st.session_state.sales_master_df[
                            st.session_state.sales_master_df['Unit Number'] == unit_no
                        ]
                        
                        if not customer_row.empty:
                            st.write(f"**Tower:** {customer_row.iloc[0].get('Tower No', 'N/A')}")
                            st.write(f"**Booking Date:** {customer_row.iloc[0].get('Booking date', 'N/A')}")
                            st.write(f"**Payment Plan:** {customer_row.iloc[0].get('Payment Plan', 'N/A')}")
                    
                    with col2:
                        st.markdown('<div class="subsection-header">Payment Verification</div>', unsafe_allow_html=True)
                        st.write(f"**Expected Amount:** ₹{verification.get('expected_amount', 0):,.2f}")
                        st.write(f"**Actual Amount:** ₹{verification.get('actual_amount', 0):,.2f}")
                        
                        difference = verification.get('expected_amount', 0) - verification.get('actual_amount', 0)
                        st.write(f"**Difference:** ₹{difference:,.2f}")
                        
                        status = verification.get('status', 'unknown')
                        if status == 'verified':
                            st.success("✅ Payments Verified")
                        elif status == 'warning':
                            st.warning("⚠️ Verification Warning")
                        else:
                            st.error("❌ Verification Failed")
                    
                    # Show transactions
                    st.markdown('<div class="subsection-header">Transactions</div>', unsafe_allow_html=True)
                    transactions = verification.get('transactions', [])
                    
                    if transactions:
                        transactions_df = pd.DataFrame(transactions)
                        
                        # Format date column
                        if 'date' in transactions_df.columns:
                            transactions_df['date'] = pd.to_datetime(transactions_df['date']).dt.strftime('%Y-%m-%d')
                        
                        # Select relevant columns
                        display_cols = ['date', 'description', 'type', 'amount', 'account_name', 'sales_tag']
                        display_cols = [col for col in display_cols if col in transactions_df.columns]
                        
                        st.dataframe(transactions_df[display_cols], use_container_width=True)
                    else:
                        st.info("No transactions found for this unit.")
                    
                    # Show bounced transactions if any
                    bounced = verification.get('bounced_transactions', [])
                    if bounced:
                        st.markdown('<div class="subsection-header">Potential Bounced Transactions</div>', unsafe_allow_html=True)
                        bounced_df = pd.DataFrame(bounced)
                        st.dataframe(bounced_df, use_container_width=True)
                    
                    # Generate preview data for this customer
                    customer_info = st.session_state.sales_master_df[
                        st.session_state.sales_master_df['Unit Number'] == unit_no
                    ].iloc[0]
                    
                    cost_sheet_data = generate_cost_sheet_data(customer_info, verification)
                    st.session_state.preview_data[unit_no] = cost_sheet_data
                    
                    # Preview button
                    if st.button(f"Preview Cost Sheet for {unit_no}", key=f"preview_{unit_no}"):
                        preview_data = st.session_state.preview_data.get(unit_no)
                        
                        if preview_data:
                            st.markdown('<div class="subsection-header">Cost Sheet Preview</div>', unsafe_allow_html=True)
                            
                            preview_tabs = st.tabs(["Data Entry", "Bank Credit Details", "Sales NOC SWAMIH"])
                            
                            with preview_tabs[0]:
                                st.write("**COST SHEET**")
                                
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.write("**Customer Information**")
                                    st.write(f"**APPLICANT NAME:** {preview_data.get('customer_name')}")
                                    st.write(f"**CO-APPLICANT NAME:** {preview_data.get('co_applicant', 'N/A')}")
                                    st.write(f"**BOOKING DATE:** {preview_data.get('booking_date')}")
                                    st.write(f"**PAYMENT PLAN:** {preview_data.get('payment_plan')}")
                                
                                with col2:
                                    st.write("**Unit Information**")
                                    st.write(f"**TOWER:** {preview_data.get('tower')}")
                                    st.write(f"**UNIT NO:** {preview_data.get('unit_number')}")
                                    st.write(f"**FLOOR NUMBER:** {preview_data.get('floor_number')}")
                                    st.write(f"**SUPER AREA:** {preview_data.get('super_area')} sq.ft.")
                                    st.write(f"**CARPET AREA:** {preview_data.get('carpet_area')} sq.ft.")
                                
                                st.write("**Cost Breakdown**")
                                cost_df = pd.DataFrame([
                                    {"Particulars": "BASIC SALE PRICE", "Rate": preview_data.get('bsp_rate'), 
                                     "Amount": preview_data.get('bsp_amount')},
                                    {"Particulars": "IFMS", "Rate": preview_data.get('ifms_rate'), 
                                     "Amount": preview_data.get('ifms_amount')},
                                    {"Particulars": "1 Year Annual Maintenance Charge", "Rate": preview_data.get('amc_rate'), 
                                     "Amount": preview_data.get('amc_amount')},
                                    {"Particulars": "Total Sale Consideration", "Rate": "", 
                                     "Amount": preview_data.get('total_consideration')},
                                    {"Particulars": "GST AMOUNT (5%)", "Rate": "", 
                                     "Amount": preview_data.get('gst_amount')},
                                    {"Particulars": "AMC GST (18%)", "Rate": "", 
                                     "Amount": preview_data.get('amc_gst_amount')},
                                    {"Particulars": "GRAND TOTAL", "Rate": "", 
                                     "Amount": preview_data.get('total_consideration') + 
                                             preview_data.get('gst_amount') + 
                                             preview_data.get('amc_gst_amount')}
                                ])
                                st.dataframe(cost_df, use_container_width=True)
                                
                                st.write("**Receipt Details**")
                                receipt_df = pd.DataFrame([
                                    {"Detail": "Amount Received", "Value": preview_data.get('amount_received')},
                                    {"Detail": "BALANCE RECEIVABLE", "Value": preview_data.get('balance_receivable')},
                                    {"Detail": "GST received", "Value": preview_data.get('gst_received')},
                                    {"Detail": "BALANCE GST RECEIVABLE", "Value": preview_data.get('gst_amount') - 
                                                                                preview_data.get('gst_received')}
                                ])
                                st.dataframe(receipt_df, use_container_width=True)
                            
                            with preview_tabs[1]:
                                st.write("**Bank Credit Details**")
                                
                                transactions = preview_data.get('transactions', [])
                                credit_transactions = [t for t in transactions if t.get('type') == 'C']
                                
                                if credit_transactions:
                                    bank_df = pd.DataFrame([
                                        {
                                            "Date": txn.get('date'),
                                            "Amount received": txn.get('amount'),
                                            "Bank A/c Name": txn.get('account_name'),
                                            "Bank A/c No": txn.get('account_number'),
                                            "Bank statement verified": "Yes"
                                        }
                                        for txn in credit_transactions
                                    ])
                                    
                                    # Add total row
                                    bank_df.loc[len(bank_df)] = {
                                        "Date": "Total Sum received",
                                        "Amount received": bank_df["Amount received"].sum(),
                                        "Bank A/c Name": "",
                                        "Bank A/c No": "",
                                        "Bank statement verified": ""
                                    }
                                    
                                    st.dataframe(bank_df, use_container_width=True)
                                else:
                                    st.info("No credit transactions found for this unit.")
                            
                            with preview_tabs[2]:
                                st.write("**Sales NOC SWAMIH**")
                                
                                noc_df = pd.DataFrame([
                                    {"Particulars": "Flat/Unit No.", "Details": preview_data.get('unit_number')},
                                    {"Particulars": "Floor No.", "Details": preview_data.get('floor_number')},
                                    {"Particulars": "Building Name", "Details": preview_data.get('tower')},
                                    {"Particulars": "Carpet Area of the Flat / Unit (Sq.Ft.)", 
                                     "Details": preview_data.get('carpet_area')},
                                    {"Particulars": "Name of the Applicant (Purchaser)", 
                                     "Details": preview_data.get('customer_name')},
                                    {"Particulars": "Name of the Co-Applicant (Co- Purchaser)", 
                                     "Details": preview_data.get('co_applicant', 'N/A')},
                                    {"Particulars": "Total Sale Consideration (including parking charges) (Excl. GST)", 
                                     "Details": preview_data.get('total_consideration')},
                                    {"Particulars": "Total GST", 
                                     "Details": preview_data.get('gst_amount') + preview_data.get('amc_gst_amount')},
                                    {"Particulars": "Sale Consideration Amount Received as on date (Excl. GST)", 
                                     "Details": preview_data.get('amount_received')},
                                    {"Particulars": "GST Amount received as on date", 
                                     "Details": preview_data.get('gst_received')},
                                    {"Particulars": "Balance Amount yet to be received (excluding taxes)", 
                                     "Details": preview_data.get('balance_receivable')},
                                    {"Particulars": "Booking Date", "Details": preview_data.get('booking_date')},
                                    {"Particulars": "Home loan taken from", "Details": preview_data.get('home_loan')},
                                ])
                                
                                st.dataframe(noc_df, use_container_width=True)
                    
            # Generate button to go to generation page
            if st.button("Generate Cost Sheets for Selected Customers", use_container_width=True):
                st.session_state.active_tab = "Generate"
                
    else:
        st.error("Please upload and process the data first.")
        if st.button("Go to Upload Page", use_container_width=True):
            st.session_state.active_tab = "Upload"

elif st.session_state.active_tab == "Dashboard":
    st.markdown('<div class="section-header">Collection Dashboard</div>', unsafe_allow_html=True)
    
    # Check if data is processed
    if st.session_state.sales_master_df is not None and st.session_state.dashboard_data:
        dashboard_data = st.session_state.dashboard_data
        
        # Show overall metrics
        st.markdown('<div class="subsection-header">Overall Collection Status</div>', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Total Units</div>
                <div class="metric-value">{dashboard_data['total_units']}</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Total Consideration</div>
                <div class="metric-value">₹{dashboard_data['total_consideration']:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Amount Received</div>
                <div class="metric-value">₹{dashboard_data['total_received']:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)
            
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Overall Completion</div>
                <div class="metric-value">{dashboard_data['overall_completion']:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Filter by collection percentage
        st.markdown('<div class="subsection-header">Filter Units by Collection Percentage</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 3])
        
        with col1:
            completion_filter = st.slider(
                "Minimum Collection Percentage", 
                min_value=0, 
                max_value=100,
                value=0,
                step=10
            )
        
        # Filter and display units
        unit_completion = dashboard_data.get('unit_completion', [])
        filtered_units = [u for u in unit_completion if u['completion_pct'] >= completion_filter]
        
        # Sort by completion percentage descending
        filtered_units.sort(key=lambda x: x['completion_pct'], reverse=True)
        
        with col2:
            st.write(f"Showing {len(filtered_units)} units with collection percentage ≥ {completion_filter}%")
            
            # Create a progress bar for each unit
            for unit in filtered_units[:10]:  # Show top 10 for performance
                col_label, col_bar = st.columns([2, 3])
                with col_label:
                    st.write(f"{unit['unit']} - {unit['customer_name']}")
                with col_bar:
                    st.progress(unit['completion_pct'] / 100)
                    st.write(f"₹{unit['amount_received']:,.0f} / ₹{unit['total_consideration']:,.0f} ({unit['completion_pct']:.1f}%)")
        
        # Create tower-wise analysis
        st.markdown('<div class="subsection-header">Tower-wise Collection Analysis</div>', unsafe_allow_html=True)
        
        tower_stats = dashboard_data.get('tower_stats', {})
        
        # Prepare data for visualization
        tower_df = pd.DataFrame([
            {
                'Tower': tower,
                'Total Units': stats['total_units'],
                'Active Units': stats['active_units'],
                'Total Consideration': stats['total_consideration'],
                'Amount Received': stats['amount_received'],
                'Completion %': (stats['amount_received'] / stats['total_consideration'] * 100) if stats['total_consideration'] > 0 else 0
            }
            for tower, stats in tower_stats.items()
        ])
        
        # Sort by total consideration
        tower_df = tower_df.sort_values('Total Consideration', ascending=False)
        
        # Show table and visualization
        col1, col2 = st.columns(2)
        
        with col1:
            st.dataframe(tower_df, use_container_width=True)
            
        with col2:
            # Create a bar chart using Plotly
            fig = px.bar(
                tower_df, 
                x='Tower', 
                y=['Amount Received', 'Total Consideration'],
                title='Collection by Tower',
                labels={'value': 'Amount (₹)', 'Tower': 'Tower', 'variable': 'Category'},
                barmode='overlay'
            )
            
            fig.update_layout(
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=400
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
        # Show collection completion distribution
        st.markdown('<div class="subsection-header">Collection Completion Distribution</div>', unsafe_allow_html=True)
        
        # Create ranges for completion percentages
        ranges = [(0, 10), (10, 25), (25, 50), (50, 75), (75, 90), (90, 100), (100, 100)]
        range_labels = ['0-10%', '10-25%', '25-50%', '50-75%', '75-90%', '90-99%', '100%']
        
        # Count units in each range
        range_counts = []
        for i, (start, end) in enumerate(ranges):
            if start == end:  # For 100% case
                count = sum(1 for u in unit_completion if u['completion_pct'] == start)
            else:
                count = sum(1 for u in unit_completion if start <= u['completion_pct'] < end)
            range_counts.append(count)
        
        # Create distribution DataFrame
        dist_df = pd.DataFrame({
            'Range': range_labels,
            'Count': range_counts
        })
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.dataframe(dist_df, use_container_width=True)
            
        with col2:
            # Create pie chart
            fig = px.pie(
                dist_df,
                values='Count',
                names='Range',
                title='Collection Completion Distribution'
            )
            
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=400)
            
            st.plotly_chart(fig, use_container_width=True)
        
        # Show all units in a table with filtering
        st.markdown('<div class="subsection-header">All Units</div>', unsafe_allow_html=True)
        
        # Create DataFrame for all units
        all_units_df = pd.DataFrame(unit_completion)
        all_units_df['completion_pct'] = all_units_df['completion_pct'].round(2)
        
        # Format columns
        all_units_df['total_consideration'] = all_units_df['total_consideration'].map('₹{:,.0f}'.format)
        all_units_df['amount_received'] = all_units_df['amount_received'].map('₹{:,.0f}'.format)
        all_units_df['completion_pct'] = all_units_df['completion_pct'].map('{:.1f}%'.format)
        
        # Rename columns for display
        all_units_df = all_units_df.rename(columns={
            'unit': 'Unit Number',
            'customer_name': 'Customer Name',
            'total_consideration': 'Total Consideration',
            'amount_received': 'Amount Received',
            'completion_pct': 'Completion %',
            'status': 'Status'
        })
        
        # Add tooltip for status
        st.write("Status Legend: ✅ Verified = Transactions match Annex data, ⚠️ Warning = Potential issues, ❌ Error = Transactions don't match")
        
        # Show dataframe with status indicators
        st.dataframe(
            all_units_df.style.applymap(
                lambda x: 'background-color: #D1FAE5' if x == 'verified' else
                         ('background-color: #FEF3C7' if x == 'warning' else
                          'background-color: #FEE2E2'),
                subset=['Status']
            ),
            use_container_width=True
        )
        
    else:
        st.error("Please upload and process the data first.")
        if st.button("Go to Upload Page", use_container_width=True):
            st.session_state.active_tab = "Upload"

elif st.session_state.active_tab == "Generate":
    st.markdown('<div class="section-header">Generate Cost Sheets</div>', unsafe_allow_html=True)
    
    # Check if customers are selected
    if not st.session_state.selected_customers:
        st.warning("No customers selected. Please go to the Customer Selection page and select at least one customer.")
        if st.button("Go to Customer Selection", use_container_width=True):
            st.session_state.active_tab = "Customers"
    else:
        st.write(f"Generating cost sheets for {len(st.session_state.selected_customers)} selected customers:")
        
        # Display selected customers
        selected_customer_info = []
        for unit_no in st.session_state.selected_customers:
            verification = st.session_state.verification_results.get(unit_no, {})
            customer_name = verification.get('customer_name', 'Unknown')
            selected_customer_info.append({
                'Unit Number': unit_no,
                'Customer Name': customer_name,
                'Status': verification.get('status', 'unknown')
            })
            
        selected_df = pd.DataFrame(selected_customer_info)
        st.dataframe(selected_df, use_container_width=True)
        
        # Generate button
        if st.button("Generate Cost Sheets and NOC Documents", use_container_width=True):
            with st.spinner(f"Generating cost sheets for {len(st.session_state.selected_customers)} customers..."):
                # Create a temporary directory to store the files
                with tempfile.TemporaryDirectory() as temp_dir:
                    cost_sheet_files = []
                    noc_files = []
                    
                    # Generate cost sheets for each selected customer
                    for unit_no in st.session_state.selected_customers:
                        # Get customer info and verification data
                        customer_row = st.session_state.sales_master_df[
                            st.session_state.sales_master_df['Unit Number'] == unit_no
                        ]
                        
                        if customer_row.empty:
                            continue
                        
                        customer_info = customer_row.iloc[0]
                        verification = st.session_state.verification_results.get(unit_no, {})
                        cost_sheet_data = generate_cost_sheet_data(customer_info, verification)
                        
                        # Generate Excel file
                        excel_file = generate_cost_sheet_excel(cost_sheet_data)
                        
                        if excel_file:
                            # Save Excel file to temp directory
                            excel_path = os.path.join(temp_dir, f"COST SHEET-{unit_no}.xlsx")
                            with open(excel_path, 'wb') as f:
                                f.write(excel_file.getvalue())
                            
                            cost_sheet_files.append((excel_path, f"COST SHEET-{unit_no}.xlsx"))
                        
                        # Generate NOC document if template is available
                        if st.session_state.noc_template:
                            noc_doc = generate_noc_document(customer_info, st.session_state.noc_template)
                            
                            if noc_doc:
                                # Save NOC file to temp directory
                                noc_path = os.path.join(temp_dir, f"NOC-{unit_no}.docx")
                                with open(noc_path, 'wb') as f:
                                    f.write(noc_doc.getvalue())
                                
                                noc_files.append((noc_path, f"NOC-{unit_no}.docx"))
                    
                    # Create a zip file if multiple files
                    if len(cost_sheet_files) > 1:
                        zip_path = os.path.join(temp_dir, "Cost_Sheets.zip")
                        with zipfile.ZipFile(zip_path, 'w') as zipf:
                            for file_path, file_name in cost_sheet_files:
                                zipf.write(file_path, file_name)
                            
                            if noc_files:
                                for file_path, file_name in noc_files:
                                    zipf.write(file_path, file_name)
                        
                        # Create download link for zip
                        with open(zip_path, 'rb') as f:
                            zip_data = f.read()
                        
                        st.success(f"Successfully generated {len(cost_sheet_files)} cost sheets and {len(noc_files)} NOC documents.")
                        st.download_button(
                            label="Download All Files (ZIP)",
                            data=zip_data,
                            file_name="Cost_Sheets.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                    else:
                        # Create individual download links
                        st.success("Cost sheet generated successfully!")
                        
                        for file_path, file_name in cost_sheet_files:
                            with open(file_path, 'rb') as f:
                                file_data = f.read()
                            
                            st.download_button(
                                label=f"Download {file_name}",
                                data=file_data,
                                file_name=file_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{file_name}",
                                use_container_width=True
                            )
                        
                        for file_path, file_name in noc_files:
                            with open(file_path, 'rb') as f:
                                file_data = f.read()
                            
                            st.download_button(
                                label=f"Download {file_name}",
                                data=file_data,
                                file_name=file_name,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                key=f"download_{file_name}",
                                use_container_width=True
                            )

# Footer
st.markdown('<div class="footer">Real Estate Cost Sheet Generator © 2025</div>', unsafe_allow_html=True)
